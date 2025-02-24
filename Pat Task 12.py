from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
import openpyxl
import time

class LoginTest:
    def __init__(self):
        # Initialize the WebDriver
        self.driver = webdriver.Chrome()
        self.url = "https://opensource-demo.orangehrmlive.com/web/index.php/auth/login"
        self.test_data = [
            {"Test ID": 1, "Username": "Admin", "Password": "admin134", "Date": "2025-02-23", "Time of Test": "10:00 AM", "Name of Tester": "Tester1", "Test Result": None},
            {"Test ID": 2, "Username": "Admin", "Password": "wrongpassword", "Date": "2025-02-23", "Time of Test": "10:10 AM", "Name of Tester": "Tester1", "Test Result": None},
            {"Test ID": 3, "Username": "Admin", "Password": "admin143", "Date": "2025-02-23", "Time of Test": "10:20 AM", "Name of Tester": "Tester2", "Test Result": None},
            {"Test ID": 4, "Username": "Admin", "Password": "wrongpassword", "Date": "2025-02-23", "Time of Test": "10:30 AM", "Name of Tester": "Tester2", "Test Result": None},
            {"Test ID": 5, "Username": "Admin", "Password": "admin123", "Date": "2025-02-23", "Time of Test": "10:40 AM", "Name of Tester": "Tester3", "Test Result": None}
        ]
        self.create_excel_file()

    def create_excel_file(self):
        # Create a new Excel workbook and sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'LoginTestData'

        # Adding headers to the Excel file
        headers = ["Test ID", "Username", "Password", "Date", "Time of Test", "Name of Tester", "Test Result"]
        sheet.append(headers)

        # Adding test data to the Excel file
        for data in self.test_data:
            sheet.append([data["Test ID"], data["Username"], data["Password"], data["Date"], data["Time of Test"], data["Name of Tester"], data["Test Result"]])

        # Save the Excel file
        wb.save("LoginTestData.xlsx")
        print("Excel file created successfully.")

    def base_url(self):
        # Navigate to the base URL and maximize the window
        self.driver.get(self.url)
        self.driver.maximize_window()

    def login(self, username, password):
        # Perform login using provided username and password
        self.base_url()
        actions = ActionChains(self.driver)
        actions.send_keys(Keys.TAB).pause(1)
        actions.send_keys(username).pause(1)
        actions.send_keys(Keys.TAB).pause(1)
        actions.send_keys(password).pause(1)
        actions.send_keys(Keys.TAB).send_keys(Keys.ENTER).pause(1)
        actions.perform()

    def check_login_status(self):
        # Check if login was successful by verifying the presence of the 'welcome' element
        try:
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.ID, "welcome"))
            )
            return True
        except:
            return False

    def update_test_result(self, test_id, result):
        # Update the test result in the Excel file
        wb = openpyxl.load_workbook("LoginTestData.xlsx")
        sheet = wb['LoginTestData']
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=7):
            if row[0].value == test_id:
                row[6].value = result
                break
        wb.save("LoginTestData.xlsx")

    def run_tests(self):
        # Iterate through the test data and perform tests
        for data in self.test_data:
            username = data["Username"]
            password = data["Password"]
            test_id = data["Test ID"]
            self.login(username, password)
            if self.check_login_status():
                print(f"Test ID {test_id} with username '{username}' PASSED.")
                self.update_test_result(test_id, "PASSED")
            else:
                print(f"Test ID {test_id} with username '{username}' FAILED.")
                self.update_test_result(test_id, "FAILED")
            self.driver.get(self.url)  # Navigate back to the login page for the next test

if __name__ == "__main__":
    # Initialize and run the tests
    task = LoginTest()
    task.run_tests()
